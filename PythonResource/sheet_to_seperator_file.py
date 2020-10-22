import openpyxl

file_path = "./Example01.xlsx"
wb = openpyxl.load_workbook(file_path)

sheets = wb.sheetnames
save_sheet_name = ""

# 추출할 시트의 인덱스를 기준으로 잡음
# 아래와 같은 경우 3
for x in range(3,len(sheets)):
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames # sheet 이름 가져오기
    save_sheet_name = wb.worksheets[x].title

    for s in sheets:
        if s != sheets[x]:
            sheet_name = wb.get_sheet_by_name(s)
            wb.remove_sheet(sheet_name) # 찾는 시트가 가져온 리스트에 없으면 리스트 상에서 삭제

    # your final wb with just Sheet1
    wb.save(str(save_sheet_name)+".xlsx")
wb.close()